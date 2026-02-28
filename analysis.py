import os
import time
from typing import Any

from fastmcp.exceptions import ToolError

from config import (
    log, mcp, FILE_TYPE_CONFIG,
    ANALYSIS_SYSTEM_PROMPT, ANALYSIS_FOLLOWUP_SYSTEM_PROMPT,
    ANALYSIS_CHECKER_SYSTEM_PROMPT,
)
from helpers import (
    extract_code, _llm_call, _truncate_output, _parse_checker_result,
    run_analysis_code, download_file,
)


async def _inspect_file(input_file_path: str, file_type: str) -> str:
    """Run a deterministic inspection script to capture file metadata.

    Returns a text summary of the file's structure (shape, dtypes, sample
    rows, null counts, sheet names, header offset detection for Excel).
    On failure, returns the error message — still useful context for the LLM.
    """
    if file_type == "csv":
        script = """
import pandas as pd

print("=== FILE INSPECTION ===")
try:
    df = pd.read_csv(input_file_path)
except Exception as e:
    print(f"ERROR reading CSV: {e}")
    # Try common alternate encodings
    try:
        df = pd.read_csv(input_file_path, encoding='latin-1')
        print("(Read successfully with latin-1 encoding)")
    except Exception as e2:
        print(f"ERROR reading CSV with latin-1: {e2}")
        raise
print(f"Shape: {df.shape[0]} rows x {df.shape[1]} columns")
print(f"\\nColumns: {df.columns.tolist()}")
print(f"\\nData types:\\n{df.dtypes.to_string()}")
nulls = df.isnull().sum()
if nulls.any():
    print(f"\\nNull counts (columns with nulls only):\\n{nulls[nulls > 0].to_string()}")
else:
    print("\\nNo null values found.")
print(f"\\nFirst 5 rows:\\n{df.head(5).to_string()}")
print(f"\\nDescribe:\\n{df.describe(include='all').to_string()}")
"""
    else:
        # Excel — also inspect raw rows with openpyxl for header offset detection
        script = """
import pandas as pd
from openpyxl import load_workbook

print("=== FILE INSPECTION ===")

# Sheet names
wb = load_workbook(input_file_path, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheet names: {sheet_names}")

# Raw row inspection (first 10 rows of first sheet) for header offset detection
ws = wb[sheet_names[0]]
print(f"\\nRaw first 10 rows of sheet '{sheet_names[0]}':")
for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
    print(f"  Row {i}: {list(row)}")
wb.close()

# Pandas view (default read)
df = pd.read_excel(input_file_path)
print(f"\\nShape (pandas default read): {df.shape[0]} rows x {df.shape[1]} columns")
print(f"\\nColumns: {df.columns.tolist()}")
print(f"\\nData types:\\n{df.dtypes.to_string()}")
nulls = df.isnull().sum()
if nulls.any():
    print(f"\\nNull counts (columns with nulls only):\\n{nulls[nulls > 0].to_string()}")
else:
    print("\\nNo null values found.")
print(f"\\nFirst 5 rows:\\n{df.head(5).to_string()}")
print(f"\\nDescribe:\\n{df.describe(include='all').to_string()}")

if len(sheet_names) > 1:
    print(f"\\nNOTE: This file has {len(sheet_names)} sheets: {sheet_names}")
    print("The above inspection shows the first sheet only. Instruct the LLM to read other sheets if needed.")
"""
    try:
        output = await run_analysis_code(script, input_file_path)
        return _truncate_output(output)
    except RuntimeError as e:
        error_msg = f"=== FILE INSPECTION FAILED ===\nError: {e}"
        log.warning(f"File inspection failed: {e}")
        return error_msg


@mcp.tool
async def analyze_file(
    file_id: str,
    instructions: str,
) -> str:
    """Analyze a CSV or Excel file using natural language instructions.

    Downloads the file from Open WebUI by file_id, runs an agentic loop
    where an LLM generates and executes pandas analysis code across multiple
    rounds, then a checker sub-agent verifies the results. Returns natural
    language analysis results followed by an audit trail of steps and code.

    Args:
        file_id: The Open WebUI file ID of the file to analyze.
        instructions: Natural language description of the analysis to perform.
    """

    # 0. Download the file and validate type
    try:
        input_file_path, file_type = await download_file(file_id)
    except RuntimeError as e:
        raise ToolError(f"Failed to download file: {e}")

    if file_type not in ("csv", "excel"):
        # Clean up before raising
        try:
            os.remove(input_file_path)
        except OSError:
            pass
        raise ToolError(
            f"File type '{FILE_TYPE_CONFIG.get(file_type, {}).get('label', file_type)}' "
            f"is not supported for analysis. Only CSV and Excel files are supported."
        )

    log.info(f"=== New analysis request: '{instructions}' | file: {input_file_path} (type: {file_type}) ===")

    max_rounds = 3
    max_error_retries = 2
    steps: list[dict[str, Any]] = []  # {"code": ..., "output": ..., "duration_s": ...}
    analysis_start = time.monotonic()
    analysis_complete_signaled = False

    try:
        # --- PRE-ANALYSIS DATA INSPECTION ---
        log.info("Running pre-analysis data inspection...")
        inspection = await _inspect_file(input_file_path, file_type)
        log.info(f"Inspection complete ({len(inspection)} chars)")
        steps.append({"code": "(automatic data inspection)", "output": inspection, "duration_s": 0})

        # --- AGENTIC ANALYSIS LOOP ---
        for round_num in range(max_rounds):
            log.info(f"--- Analysis round {round_num + 1}/{max_rounds} ---")

            if round_num == 0:
                # First round: fresh code generation with inspection context
                messages = [
                    {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
                    {"role": "user", "content": (
                        f"Analyze this file according to these instructions: {instructions}\n\n"
                        f"DATA INSPECTION:\n{inspection}"
                    )},
                ]
            else:
                # Subsequent rounds: include prior steps, ask LLM to continue or stop
                steps_text = ""
                for i, step in enumerate(steps[1:], 1):  # skip inspection step
                    steps_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

                messages = [
                    {"role": "system", "content": ANALYSIS_FOLLOWUP_SYSTEM_PROMPT},
                    {"role": "user", "content": (
                        f"Original instructions: {instructions}\n\n"
                        f"Here are the analysis steps completed so far:\n{steps_text}\n\n"
                        f"Should more analysis be done, or is this complete?"
                    )},
                ]

            # Generate code (or completion signal)
            log.info(f"Calling LLM for analysis round {round_num + 1}...")
            raw = await _llm_call(messages)
            log.debug(f"LLM response (round {round_num + 1}):\n{raw}")

            # Check for completion signal
            if "__ANALYSIS_COMPLETE__" in raw:
                log.info(f"LLM signaled analysis complete at round {round_num + 1}")
                analysis_complete_signaled = True
                break

            code = extract_code(raw)

            # Execute with error retries
            output = None
            last_error = ""
            for retry in range(max_error_retries + 1):
                try:
                    t0 = time.monotonic()
                    output = await run_analysis_code(code, input_file_path)
                    break
                except RuntimeError as e:
                    last_error = str(e)
                    log.warning(f"Analysis code failed (retry {retry + 1}): {last_error}")
                    if retry < max_error_retries:
                        # Ask LLM to fix the code
                        fix_messages = [
                            {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
                            {"role": "user", "content": (
                                f"The previous analysis code failed.\n\n"
                                f"USER INSTRUCTIONS: {instructions}\n\n"
                                f"DATA INSPECTION:\n{inspection}\n\n"
                                f"FAILED CODE:\n{code}\n\n"
                                f"ERROR MESSAGE:\n{last_error}\n\n"
                                f"Rewrite the code to fix this error. "
                                f"Ensure you still use `input_file_path`."
                            )},
                        ]
                        raw_fix = await _llm_call(fix_messages)
                        code = extract_code(raw_fix)

            if output is None:
                log.error(f"Round {round_num + 1}: all error retries exhausted. Last error: {last_error}")
                raise ToolError(
                    f"Analysis code failed after {max_error_retries + 1} attempts. "
                    f"Last error: {last_error}"
                )

            duration = time.monotonic() - t0
            steps.append({"code": code, "output": output, "duration_s": round(duration, 1)})
            log.info(f"Round {round_num + 1} complete. Output length: {len(output)} chars, duration: {duration:.1f}s")

        # Check for analysis content beyond inspection
        analysis_steps = [s for s in steps if s["code"] != "(automatic data inspection)"]
        if not analysis_steps:
            raise ToolError("Analysis produced no results. The LLM did not generate any code.")

        if not analysis_complete_signaled:
            log.warning(f"Analysis reached maximum rounds ({max_rounds}) without explicit completion signal")

        # --- SUB-AGENT CHECKER ---
        checker_status, analysis_result = await _run_checker(instructions, analysis_steps)
        checker_note = "PASSED"
        checker_caveat = ""

        # If checker failed, redo once with feedback
        if checker_status == "failed":
            log.info(f"Checker flagged issues, running one redo. Feedback: {analysis_result[:200]}")
            checker_feedback = analysis_result

            redo_steps = await _run_analysis_redo(instructions, checker_feedback, input_file_path, steps)
            steps.extend(redo_steps)

            # Re-check (accept whatever comes out)
            redo_status, analysis_result = await _run_checker(instructions, [s for s in steps if s["code"] != "(automatic data inspection)"])
            if redo_status == "failed":
                # Accept anyway on second pass
                checker_note = "ACCEPTED WITH CAVEATS"
                checker_caveat = f"\n  - Checker feedback: {analysis_result[:300]}"
                # Use the redo result as the summary if it's not just an error message
                if not analysis_result.startswith("EXECUTION FAILED"):
                    pass  # analysis_result is already set from checker
            else:
                checker_note = "PASSED (after correction)"

        # --- FORMAT OUTPUT ---
        total_time = time.monotonic() - analysis_start
        analysis_rounds = len([s for s in steps if s["code"] != "(automatic data inspection)"])
        redo_happened = checker_note != "PASSED"

        qa_section = (
            f"## Quality Assurance\n"
            f"- Checker: {checker_note}{checker_caveat}\n"
            f"- Analysis rounds: {analysis_rounds}{' (+ redo)' if redo_happened else ''}\n"
            f"- Total execution time: {total_time:.1f}s\n"
        )
        if not analysis_complete_signaled:
            qa_section += f"- Note: Analysis reached maximum rounds ({max_rounds}) without explicit completion signal\n"

        methodology = ""
        for i, step in enumerate(steps, 0):
            if step["code"] == "(automatic data inspection)":
                methodology += f"\n### Data Inspection\n**Output:**\n```\n{step['output']}\n```\n"
            else:
                duration_note = f" ({step.get('duration_s', '?')}s)" if step.get("duration_s") else ""
                methodology += (
                    f"\n### Step {i}{duration_note}\n"
                    f"**Code:**\n```python\n{step['code']}\n```\n"
                    f"**Output:**\n```\n{step['output']}\n```\n"
                )

        return (
            f"## Analysis Results\n\n{analysis_result}\n\n"
            f"---\n\n{qa_section}\n"
            f"---\n\n## Methodology\n{methodology}"
        )

    finally:
        try:
            os.remove(input_file_path)
        except OSError:
            pass


async def _run_checker(instructions: str, steps: list[dict[str, str]]) -> tuple[str, str]:
    """Run the sub-agent checker on analysis results.

    Returns (status, content) where status is 'passed', 'failed', or 'unknown'.
    """
    steps_text = ""
    for i, step in enumerate(steps, 1):
        steps_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

    messages = [
        {"role": "system", "content": ANALYSIS_CHECKER_SYSTEM_PROMPT},
        {"role": "user", "content": (
            f"Original instructions: {instructions}\n\n"
            f"Analysis steps:\n{steps_text}"
        )},
    ]
    log.info("Running analysis checker sub-agent...")
    raw = await _llm_call(messages)
    status, content = _parse_checker_result(raw)
    log.info(f"Checker status: {status}")
    log.debug(f"Checker content:\n{content[:500]}")
    return status, content


async def _run_analysis_redo(
    instructions: str,
    checker_feedback: str,
    input_file_path: str,
    prior_steps: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Re-run analysis with checker feedback. Returns new steps.

    On execution failure, records the failed attempt in the returned steps
    (with error message as output) instead of silently dropping it.
    """
    prior_text = ""
    for i, step in enumerate(prior_steps, 1):
        prior_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

    messages = [
        {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
        {"role": "user", "content": (
            f"Analyze this file according to these instructions: {instructions}\n\n"
            f"A reviewer found these issues with the previous analysis:\n{checker_feedback}\n\n"
            f"Previous analysis steps for context:\n{prior_text}\n\n"
            f"Write corrected analysis code that addresses the reviewer's feedback."
        )},
    ]
    log.info("Running analysis redo based on checker feedback...")
    raw = await _llm_call(messages)
    code = extract_code(raw)

    new_steps = []
    try:
        t0 = time.monotonic()
        output = await run_analysis_code(code, input_file_path)
        duration = time.monotonic() - t0
        new_steps.append({"code": code, "output": output, "duration_s": round(duration, 1)})
    except RuntimeError as e:
        log.error(f"Redo analysis code failed: {e}")
        new_steps.append({"code": code, "output": f"EXECUTION FAILED: {e}", "duration_s": 0})

    return new_steps
