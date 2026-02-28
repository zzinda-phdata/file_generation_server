import os
import sys
import asyncio
import uuid
import logging
import re
import tempfile
import time
from typing import Any
import aiohttp
import uvicorn

from fastmcp import FastMCP
from fastmcp.exceptions import ToolError
from starlette.requests import Request
from starlette.responses import JSONResponse
from openai import AsyncOpenAI, APIError, APIConnectionError, RateLimitError

# --- LOGGING ---
log = logging.getLogger("openwebui_fileagent")
log.setLevel(logging.DEBUG)
_handler = logging.StreamHandler(sys.stderr)
_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"))
log.addHandler(_handler)
log.propagate = False

# --- CONFIGURATION ---
mcp = FastMCP("FileAgent")

# Temporary directory for generated files
STORAGE_DIR = tempfile.mkdtemp(prefix="owui_files_")
SCRIPT_TIMEOUT = int(os.getenv("SCRIPT_TIMEOUT", "90"))
MAX_ANALYSIS_OUTPUT_CHARS = 30_000
# These fire at module import — once per gunicorn worker when running multi-worker
log.info(f"STORAGE_DIR = {STORAGE_DIR}")
log.info(f"SCRIPT_TIMEOUT = {SCRIPT_TIMEOUT}")
log.info(f"Python executable = {sys.executable}")

client = AsyncOpenAI(
    api_key=os.getenv("OPENAI_API_KEY", "sk-placeholder"),
    base_url=os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1"),
    timeout=120.0,
)
MODEL_NAME = os.getenv("MODEL_NAME", "gpt-4o")
DEEP_RESEARCH_MODEL = os.getenv("DEEP_RESEARCH_MODEL", "o4-mini-deep-research")
DEEP_RESEARCH_POLL_INTERVAL = 5   # seconds between status checks
DEEP_RESEARCH_TIMEOUT = 600       # max seconds to wait for completion

# --- SYSTEM PROMPTS ---
EXCEL_SYSTEM_PROMPT = """
You are an expert Python developer specialized in the 'xlsxwriter' library.
Your goal is to write a Python script that generates an Excel file based on user instructions.

RULES:
1. You must use 'xlsxwriter' to create the file.
2. A variable named `file_path` will be provided to your script at runtime. You MUST use this variable when opening the workbook:
   `workbook = xlsxwriter.Workbook(file_path)`
3. You must explicitly close the workbook at the end: `workbook.close()`
4. Do NOT output markdown formatting (like ```python). Output ONLY raw Python code.
5. If you use calculations, prefer Excel formulas (e.g., `write_formula`) over pre-calculated values unless asked otherwise.
6. Make the Excel file professional: use bold headers, currency formats, and colors where appropriate.
"""

DOCX_SYSTEM_PROMPT = """
You are an expert Python developer specialized in the 'python-docx' library.
Your goal is to write a Python script that generates a Word document based on user instructions.

RULES:
1. You must use 'python-docx' (import docx / from docx import Document) to create the file.
2. A variable named `file_path` will be provided to your script at runtime. You MUST save the document to this path:
   `document.save(file_path)`
3. Do NOT output markdown formatting (like ```python). Output ONLY raw Python code.
4. Make the document professional: use appropriate headings, styles, tables, and formatting where appropriate.
"""

EXCEL_MODIFY_SYSTEM_PROMPT = """
You are an expert Python developer specialized in the 'openpyxl' library.
Your goal is to write a Python script that reads an existing Excel file and modifies it based on user instructions.

RULES:
1. You must use 'openpyxl' to read and modify the file.
2. Two variables will be provided at runtime:
   - `input_file_path` — the path to the existing Excel file (read from this)
   - `file_path` — the path where the modified file must be saved (write to this)
3. Open the existing workbook like this:
   `from openpyxl import load_workbook`
   `wb = load_workbook(input_file_path)`
4. After making modifications, save to the output path:
   `wb.save(file_path)`
5. Do NOT overwrite `input_file_path`. Do NOT save to `input_file_path`.
6. Do NOT output markdown formatting (like ```python). Output ONLY raw Python code.
7. Do NOT redefine `input_file_path` or `file_path`.
8. If the user asks to add formatting, use openpyxl styles (Font, PatternFill, Alignment, Border, etc.).
9. If the user asks to add new data, append it without destroying existing data unless explicitly told to replace.
10. Preserve existing formatting and data unless the user explicitly asks to change or remove it.
"""

DOCX_MODIFY_SYSTEM_PROMPT = """
You are an expert Python developer specialized in the 'python-docx' library.
Your goal is to write a Python script that reads an existing Word document and modifies it based on user instructions.

RULES:
1. You must use 'python-docx' (import docx / from docx import Document) to read and modify the file.
2. Two variables will be provided at runtime:
   - `input_file_path` — the path to the existing Word document (read from this)
   - `file_path` — the path where the modified document must be saved (write to this)
3. Open the existing document like this:
   `from docx import Document`
   `document = Document(input_file_path)`
4. After making modifications, save to the output path:
   `document.save(file_path)`
5. Do NOT overwrite `input_file_path`. Do NOT save to `input_file_path`.
6. Do NOT output markdown formatting (like ```python). Output ONLY raw Python code.
7. Do NOT redefine `input_file_path` or `file_path`.
8. Preserve existing content and formatting unless the user explicitly asks to change or remove it.
"""

ANALYSIS_SYSTEM_PROMPT = """
You are an expert Python data analyst. Your goal is to write a Python script using pandas to analyze a data file and print the results.

RULES:
1. A variable named `input_file_path` will be provided at runtime. Use it to read the file.
2. Detect the file type from the extension:
   - `.csv`: use `pd.read_csv(input_file_path)`
   - `.xlsx` or `.xls`: use `pd.read_excel(input_file_path)`
3. Print ALL analysis results to stdout using `print()`. This is how your results are captured.
4. Do NOT redefine `input_file_path`.
5. Do NOT output markdown formatting (like ```python). Output ONLY raw Python code.
6. Do NOT create or save any files. Only print results.
7. Use pandas, and standard library modules only.
8. Format numerical output clearly (e.g., round floats, align columns where helpful).
9. You will receive a DATA INSPECTION section showing the file's structure, columns, types,
   and sample rows. Use this to write correct code on the first attempt. Pay attention to:
   - Whether the data starts on a row other than 1 (header offset — use skiprows if needed)
   - Column data types (especially object columns that should be numeric — use pd.to_numeric with errors='coerce')
   - Missing values (NaN counts per column)
   - For Excel files with multiple sheets: which sheet(s) contain the relevant data
"""

ANALYSIS_FOLLOWUP_SYSTEM_PROMPT = """
You are an expert Python data analyst continuing a multi-step analysis.

You will receive the code and output from previous analysis steps. Based on these results and the user's original instructions, decide what to do:

OPTION A — If the analysis is COMPLETE and fully addresses the user's instructions:
  Output ONLY the text: __ANALYSIS_COMPLETE__
  Do NOT output any code.

OPTION B — If more analysis is needed:
  Output ONLY raw Python code (no markdown fences) for the next analysis step.
  The variable `input_file_path` is already defined. Do NOT redefine it.
  Print all results to stdout using `print()`.
  Do NOT create or save any files.
  Use pandas, and standard library modules only.

Choose Option A when the previous output already contains enough information to fully answer the user's question. Do not run unnecessary extra steps.
"""

ANALYSIS_CHECKER_SYSTEM_PROMPT = """
You are a senior data analyst reviewing another analyst's work. You will receive:
- The original analysis instructions
- All code that was executed (each step)
- All output produced (each step)

Your job is to verify the analysis is correct and complete.

RULES:
1. Check that the analysis actually addresses the user's instructions.
2. Check for logical errors, misinterpretations, or incorrect calculations.
3. Check that conclusions are supported by the data shown in the output.

If the analysis is CORRECT and COMPLETE:
  Start your response with: __CHECK_PASSED__
  Then write a polished, natural language summary of the analysis results.
  This summary should be clear, well-organized, and directly answer the user's question.
  Do NOT include code in this summary.

If there are SIGNIFICANT issues (wrong answers, missing key parts of the instructions, logical errors):
  Start your response with: __CHECK_FAILED__
  Then describe the specific issues that need to be fixed.
  Be concrete: say what is wrong and what the correct approach should be.

Minor formatting issues or style preferences are NOT grounds for failure. Only fail for substantive errors.
"""

# --- FILE TYPE CONFIG ---
FILE_TYPE_CONFIG: dict[str, dict[str, Any]] = {
    "excel": {"ext": ".xlsx", "system_prompt": EXCEL_SYSTEM_PROMPT, "modify_system_prompt": EXCEL_MODIFY_SYSTEM_PROMPT, "label": "Excel"},
    "docx":  {"ext": ".docx", "system_prompt": DOCX_SYSTEM_PROMPT, "modify_system_prompt": DOCX_MODIFY_SYSTEM_PROMPT, "label": "Word"},
    "csv":   {"ext": ".csv", "label": "CSV"},
}

# --- HELPER FUNCTIONS ---

def extract_code(llm_response: str) -> str:
    """Clean markdown code blocks if the LLM ignores instructions."""
    pattern = r"```(?:python)?(.*?)```"
    match = re.search(pattern, llm_response, re.DOTALL)
    if match:
        return match.group(1).strip()
    return llm_response.strip()

async def _llm_call(messages: list[dict], temperature: float = 0.01) -> str:
    """Call the LLM with retry logic for transient failures.

    Retries up to 2 times with exponential backoff on APIError,
    APIConnectionError, and RateLimitError.  Returns the response
    content string.
    """
    max_retries = 2
    for attempt in range(max_retries + 1):
        try:
            response = await client.chat.completions.create(
                model=MODEL_NAME,
                messages=messages,
                temperature=temperature,
            )
            return response.choices[0].message.content
        except (APIError, APIConnectionError, RateLimitError) as e:
            if attempt < max_retries:
                wait = [1, 3][attempt]
                log.warning(f"LLM call failed (attempt {attempt + 1}/{max_retries + 1}), retrying in {wait}s: {e}")
                await asyncio.sleep(wait)
            else:
                log.error(f"LLM call failed after {max_retries + 1} attempts: {e}")
                raise


def _truncate_output(text: str, max_chars: int = MAX_ANALYSIS_OUTPUT_CHARS) -> str:
    """Truncate text for LLM context, preserving start and noting truncation."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + f"\n... [truncated, showing first {max_chars:,} of {len(text):,} chars]"


async def _inspect_file(input_file_path: str, file_type: str) -> str:
    """Run a deterministic inspection script to capture file metadata.

    Returns a text summary of the file's structure (shape, dtypes, sample
    rows, null counts, sheet names, header offset detection for Excel).
    On failure, returns the error message — still useful context for the LLM.
    """
    if file_type == "csv":
        script = """
import pandas as pd

print("=== FILE INSPECTION ===")
try:
    df = pd.read_csv(input_file_path)
except Exception as e:
    print(f"ERROR reading CSV: {e}")
    # Try common alternate encodings
    try:
        df = pd.read_csv(input_file_path, encoding='latin-1')
        print("(Read successfully with latin-1 encoding)")
    except Exception as e2:
        print(f"ERROR reading CSV with latin-1: {e2}")
        raise
print(f"Shape: {df.shape[0]} rows x {df.shape[1]} columns")
print(f"\\nColumns: {df.columns.tolist()}")
print(f"\\nData types:\\n{df.dtypes.to_string()}")
nulls = df.isnull().sum()
if nulls.any():
    print(f"\\nNull counts (columns with nulls only):\\n{nulls[nulls > 0].to_string()}")
else:
    print("\\nNo null values found.")
print(f"\\nFirst 5 rows:\\n{df.head(5).to_string()}")
print(f"\\nDescribe:\\n{df.describe(include='all').to_string()}")
"""
    else:
        # Excel — also inspect raw rows with openpyxl for header offset detection
        script = """
import pandas as pd
from openpyxl import load_workbook

print("=== FILE INSPECTION ===")

# Sheet names
wb = load_workbook(input_file_path, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheet names: {sheet_names}")

# Raw row inspection (first 10 rows of first sheet) for header offset detection
ws = wb[sheet_names[0]]
print(f"\\nRaw first 10 rows of sheet '{sheet_names[0]}':")
for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
    print(f"  Row {i}: {list(row)}")
wb.close()

# Pandas view (default read)
df = pd.read_excel(input_file_path)
print(f"\\nShape (pandas default read): {df.shape[0]} rows x {df.shape[1]} columns")
print(f"\\nColumns: {df.columns.tolist()}")
print(f"\\nData types:\\n{df.dtypes.to_string()}")
nulls = df.isnull().sum()
if nulls.any():
    print(f"\\nNull counts (columns with nulls only):\\n{nulls[nulls > 0].to_string()}")
else:
    print("\\nNo null values found.")
print(f"\\nFirst 5 rows:\\n{df.head(5).to_string()}")
print(f"\\nDescribe:\\n{df.describe(include='all').to_string()}")

if len(sheet_names) > 1:
    print(f"\\nNOTE: This file has {len(sheet_names)} sheets: {sheet_names}")
    print("The above inspection shows the first sheet only. Instruct the LLM to read other sheets if needed.")
"""
    try:
        output = await run_analysis_code(script, input_file_path)
        return _truncate_output(output)
    except RuntimeError as e:
        error_msg = f"=== FILE INSPECTION FAILED ===\nError: {e}"
        log.warning(f"File inspection failed: {e}")
        return error_msg


def _parse_checker_result(raw: str) -> tuple[str, str]:
    """Parse checker LLM response into (status, content).

    Status is 'passed', 'failed', or 'unknown'.
    Content is the response with the marker stripped.
    """
    stripped = raw.strip()
    if stripped.startswith("__CHECK_PASSED__"):
        return "passed", stripped[len("__CHECK_PASSED__"):].strip()
    elif stripped.startswith("__CHECK_FAILED__"):
        return "failed", stripped[len("__CHECK_FAILED__"):].strip()
    else:
        log.warning("Checker response missing expected marker, treating as passed")
        return "unknown", stripped


async def generate_code(instructions: str, system_prompt: str, error_context: str = None, previous_code: str = None, input_file_path: str = None) -> str:
    """Call the LLM to generate Python code, or to fix previously failed code.

    On the first attempt, sends user instructions directly.  On retries,
    includes the failed code and error message so the LLM can self-correct.
    When *input_file_path* is provided (modification flow), prompts reference
    both variables.
    """
    messages = [{"role": "system", "content": system_prompt}]

    if error_context:
        log.debug(f"Requesting code fix — error was: {error_context}")
        if input_file_path:
            path_reminder = (
                "Ensure you still read from `input_file_path` and "
                "save to `file_path`. Do NOT redefine these variables."
            )
        else:
            path_reminder = "Ensure you still use `file_path`."
        messages.append({
            "role": "user",
            "content": f"""
            The previous code you wrote failed.

            USER INSTRUCTIONS: {instructions}

            FAILED CODE:
            {previous_code}

            ERROR MESSAGE:
            {error_context}

            Please rewrite the code to fix this error. {path_reminder}
            """
        })
    else:
        log.debug(f"Requesting fresh code generation for: {instructions}")
        if input_file_path:
            messages.append({
                "role": "user",
                "content": (
                    f"Modify an existing file according to these instructions: {instructions}\n\n"
                    f"The existing file is available at the path stored in `input_file_path`. "
                    f"Save the modified file to the path stored in `file_path`."
                )
            })
        else:
            messages.append({"role": "user", "content": f"Create a python script for this request: {instructions}"})

    log.info(f"Calling LLM ({MODEL_NAME})...")
    response = await client.chat.completions.create(
        model=MODEL_NAME,
        messages=messages,
        temperature=0.01
    )
    raw = response.choices[0].message.content
    code = extract_code(raw)
    log.debug(f"Generated code:\n{code}")
    return code

async def run_code_in_subprocess(code: str, file_path: str, input_file_path: str = None) -> None:
    """Run LLM-generated code in an isolated subprocess.

    The subprocess inherits only PATH (no API keys or secrets).  Any
    ``file_path = ...`` assignments the LLM may have generated are stripped
    and replaced with the correct runtime path.  When *input_file_path* is
    provided (modification flow), it is also injected and protected.
    """
    # Strip any file_path reassignments the LLM may have generated
    cleaned = re.sub(r"(?m)^file_path\s*=\s*.+$", "", code)
    if input_file_path:
        cleaned = re.sub(r"(?m)^input_file_path\s*=\s*.+$", "", cleaned)
        script_content = (
            f"input_file_path = {input_file_path!r}\n"
            f"file_path = {file_path!r}\n"
            f"{cleaned}"
        )
    else:
        script_content = f"file_path = {file_path!r}\n{cleaned}"
    script_path = os.path.join(STORAGE_DIR, f"_script_{uuid.uuid4().hex[:8]}.py")
    log.info(f"Writing script to {script_path}")
    log.debug(f"Script content:\n{script_content}")
    try:
        with open(script_path, "w") as f:
            f.write(script_content)

        safe_env = {"PATH": os.getenv("PATH", "/usr/bin:/usr/local/bin")}
        log.info(f"Running subprocess: {sys.executable} {script_path}")
        proc = await asyncio.create_subprocess_exec(
            sys.executable, script_path,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env=safe_env,
        )
        try:
            stdout, stderr = await asyncio.wait_for(
                proc.communicate(), timeout=SCRIPT_TIMEOUT
            )
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            raise RuntimeError(f"Script timed out after {SCRIPT_TIMEOUT} seconds")

        stdout_text = stdout.decode() if stdout else ""
        stderr_text = stderr.decode() if stderr else ""

        log.info(f"Subprocess exited with code {proc.returncode}")
        if stdout_text:
            log.info(f"[stdout] {stdout_text.strip()}")
        if stderr_text:
            log.warning(f"[stderr] {stderr_text.strip()}")
        if proc.returncode != 0:
            raise RuntimeError(stderr_text.strip() or f"Script exited with code {proc.returncode}")

        log.info(f"Checking for output file at: {file_path}")
        log.info(f"File exists: {os.path.exists(file_path)}")
        if os.path.exists(file_path):
            log.info(f"File size: {os.path.getsize(file_path)} bytes")
    finally:
        try:
            os.remove(script_path)
        except OSError:
            pass


async def run_analysis_code(code: str, input_file_path: str) -> str:
    """Run pandas analysis code in an isolated subprocess and return stdout.

    Similar to run_code_in_subprocess but returns captured stdout text
    instead of writing to an output file.  Raises RuntimeError on failure.
    """
    cleaned = re.sub(r"(?m)^input_file_path\s*=\s*.+$", "", code)
    script_content = f"input_file_path = {input_file_path!r}\n{cleaned}"
    script_path = os.path.join(STORAGE_DIR, f"_analysis_{uuid.uuid4().hex[:8]}.py")
    log.info(f"Writing analysis script to {script_path}")
    log.debug(f"Analysis script content:\n{script_content}")
    try:
        with open(script_path, "w") as f:
            f.write(script_content)

        safe_env = {"PATH": os.getenv("PATH", "/usr/bin:/usr/local/bin")}
        proc = await asyncio.create_subprocess_exec(
            sys.executable, script_path,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env=safe_env,
        )
        try:
            stdout, stderr = await asyncio.wait_for(
                proc.communicate(), timeout=SCRIPT_TIMEOUT
            )
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            raise RuntimeError(f"Analysis script timed out after {SCRIPT_TIMEOUT} seconds")

        stdout_text = stdout.decode() if stdout else ""
        stderr_text = stderr.decode() if stderr else ""

        log.info(f"Analysis subprocess exited with code {proc.returncode}")
        if stdout_text:
            log.debug(f"[analysis stdout] {stdout_text[:500]}")
        if proc.returncode != 0:
            raise RuntimeError(stderr_text.strip() or f"Script exited with code {proc.returncode}")

        # Append meaningful stderr warnings (filter out noise)
        if stderr_text:
            _noise = ("FutureWarning", "DeprecationWarning", "PendingDeprecationWarning")
            warning_lines = [
                line for line in stderr_text.strip().splitlines()
                if not any(noise in line for noise in _noise)
            ]
            if warning_lines:
                stdout_text += "\n--- Warnings ---\n" + "\n".join(warning_lines)

        return stdout_text
    finally:
        try:
            os.remove(script_path)
        except OSError:
            pass


async def upload_file(local_path: str, filename: str) -> str:
    """Upload a generated file to Open WebUI and return a public download URL.

    Raises on failure so the retry loop can catch and self-correct.
    """
    internal_api_url = os.getenv("INTERNAL_API_URL", "http://localhost:8080")
    public_domain = os.getenv("PUBLIC_DOMAIN", "http://localhost:8080")
    api_key = os.getenv("OPENWEBUI_API_KEY", "sk-placeholder")
    headers = {"Authorization": f"Bearer {api_key}"}
    upload_url = f"{internal_api_url.rstrip('/')}/api/v1/files/"

    async with aiohttp.ClientSession(
        timeout=aiohttp.ClientTimeout(total=60)
    ) as session:
        with open(local_path, "rb") as f:
            form = aiohttp.FormData()
            form.add_field("file", f, filename=filename)
            async with session.post(upload_url, headers=headers, data=form) as resp:
                status = resp.status
                text = await resp.text()
                try:
                    data = await resp.json()
                except Exception:
                    data = {}

    try:
        os.remove(local_path)
    except Exception:
        pass

    if status < 200 or status >= 300:
        raise RuntimeError(f"Upload failed (HTTP {status}): {text}")

    file_id = data.get("id") or data.get("uuid") or data.get("file_id")
    return f"{public_domain.rstrip('/')}/api/v1/files/{file_id}/content"


# Extension-to-file-type mapping (reverse of FILE_TYPE_CONFIG)
_EXT_TO_FILE_TYPE = {".xlsx": "excel", ".xls": "excel", ".docx": "docx", ".csv": "csv"}

# Content-Type to file type mapping
_CONTENT_TYPE_TO_FILE_TYPE = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "excel",
    "application/vnd.ms-excel": "excel",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "text/csv": "csv",
    "application/csv": "csv",
}


def _detect_file_type(resp: aiohttp.ClientResponse) -> str:
    """Detect file type from HTTP response headers.

    Checks Content-Disposition filename extension first, then Content-Type.
    Raises RuntimeError if the file type cannot be determined or is unsupported.
    """
    # 1. Try Content-Disposition header for filename
    cd = resp.headers.get("Content-Disposition", "")
    if cd:
        match = re.search(r'filename[*]?=["\']?([^"\';\s]+)', cd)
        if match:
            filename = match.group(1)
            ext = os.path.splitext(filename)[1].lower()
            if ext in _EXT_TO_FILE_TYPE:
                return _EXT_TO_FILE_TYPE[ext]
            log.warning(f"Content-Disposition filename '{filename}' has unrecognized extension '{ext}'")

    # 2. Try Content-Type header
    content_type = resp.content_type or ""
    base_type = content_type.split(";")[0].strip().lower()
    if base_type in _CONTENT_TYPE_TO_FILE_TYPE:
        return _CONTENT_TYPE_TO_FILE_TYPE[base_type]

    # 3. Cannot determine
    raise RuntimeError(
        f"Cannot determine file type from response headers. "
        f"Content-Disposition: '{cd}', Content-Type: '{content_type}'. "
        f"Supported types: {', '.join(FILE_TYPE_CONFIG.keys())}"
    )


async def download_file(file_id: str) -> tuple[str, str]:
    """Download a file from Open WebUI by file_id.

    Returns (local_path, file_type) where file_type is a key in FILE_TYPE_CONFIG.
    Raises RuntimeError on download failure or unrecognized file type.
    """
    internal_api_url = os.getenv("INTERNAL_API_URL", "http://localhost:8080")
    api_key = os.getenv("OPENWEBUI_API_KEY", "sk-placeholder")
    headers = {"Authorization": f"Bearer {api_key}"}
    download_url = f"{internal_api_url.rstrip('/')}/api/v1/files/{file_id}/content"

    log.info(f"Downloading file {file_id} from {download_url}")

    async with aiohttp.ClientSession(
        timeout=aiohttp.ClientTimeout(total=120)
    ) as session:
        async with session.get(download_url, headers=headers) as resp:
            if resp.status < 200 or resp.status >= 300:
                text = await resp.text()
                raise RuntimeError(
                    f"Failed to download file {file_id} (HTTP {resp.status}): {text}"
                )

            file_type = _detect_file_type(resp)

            local_filename = f"input_{file_id}_{uuid.uuid4().hex[:8]}{FILE_TYPE_CONFIG[file_type]['ext']}"
            local_path = os.path.join(STORAGE_DIR, local_filename)

            with open(local_path, "wb") as f:
                async for chunk in resp.content.iter_chunked(8192):
                    f.write(chunk)

    log.info(f"Downloaded file to {local_path} (type: {file_type}, size: {os.path.getsize(local_path)} bytes)")
    return local_path, file_type


# --- MCP TOOL ---

@mcp.tool
async def generate_file(
    instructions: str,
    file_type: str = "excel",
    filename_hint: str = "output",
) -> str:
    """Generate an Excel or Word file from natural language instructions.

    Takes user instructions and a file_type, generates Python code to build
    the file, executes it in a sandbox, uploads to Open WebUI, and returns
    a download link. Self-corrects if code generation fails.

    Args:
        instructions: Natural language description of the file to generate.
        file_type: File type to generate: 'excel' or 'docx'.
        filename_hint: Base name for the generated file (UUID suffix added automatically).
    """

    # 0. Validate file_type
    if file_type not in FILE_TYPE_CONFIG or "system_prompt" not in FILE_TYPE_CONFIG.get(file_type, {}):
        supported = [k for k, v in FILE_TYPE_CONFIG.items() if "system_prompt" in v]
        raise ToolError(
            f"Unsupported file_type '{file_type}'. Must be one of: {', '.join(supported)}"
        )
    config = FILE_TYPE_CONFIG[file_type]

    # 1. Setup paths
    filename = f"{filename_hint}_{uuid.uuid4().hex[:8]}{config['ext']}"
    file_path = os.path.join(STORAGE_DIR, filename)

    # 2. The Healing Loop
    max_attempts = 4
    current_code = ""
    last_error = ""

    log.info(f"=== New {config['label']} request: '{instructions}' | file: {filename} ===")
    log.info(f"Target path: {file_path}")

    for attempt in range(max_attempts):
        try:
            # A. Generate Code (Fresh or Fix)
            if attempt == 0:
                log.info(f"--- Attempt {attempt + 1}/{max_attempts}: fresh generation ---")
                current_code = await generate_code(instructions, config["system_prompt"])
            else:
                log.info(f"--- Attempt {attempt + 1}/{max_attempts}: self-correction (error: {last_error}) ---")
                current_code = await generate_code(instructions, config["system_prompt"], last_error, current_code)

            # B. Execute in isolated subprocess
            await run_code_in_subprocess(current_code, file_path)

            # C. Verify File Creation
            if not os.path.exists(file_path):
                raise Exception(
                    f"Code ran without error, but no file was found at '{file_path}'. "
                    "Did you forget to save/close the file?"
                )

            # D. Upload to Open WebUI
            log.info(f"Uploading {filename} to Open WebUI...")
            download_url = await upload_file(file_path, filename)

            log.info(f"Success on attempt {attempt + 1}: {download_url}")
            return (
                f"{config['label']} file generated successfully "
                f"(attempt {attempt + 1} of {max_attempts}).\n"
                f"Download: {download_url}"
            )

        except ToolError:
            raise
        except Exception as e:
            last_error = str(e)
            log.error(f"Attempt {attempt + 1} failed: {last_error}", exc_info=True)
            # Clean up partial output file before retrying
            try:
                os.remove(file_path)
            except OSError:
                pass

    # 3. Failure after all attempts
    log.error(f"All {max_attempts} attempts exhausted. Last error: {last_error}")
    raise ToolError(
        f"Failed to generate valid {config['label']} file after {max_attempts} attempts. "
        f"Last error: {last_error}"
    )


@mcp.tool
async def modify_file(
    file_id: str,
    instructions: str,
    filename_hint: str = "modified",
) -> str:
    """Modify an existing Excel or Word file using natural language instructions.

    Downloads the file from Open WebUI by file_id, uses an LLM to generate
    Python code that reads and modifies it, executes the code in a sandbox,
    uploads the result back to Open WebUI, and returns a download link.

    Args:
        file_id: The Open WebUI file ID of the file to modify.
        instructions: Natural language description of the modifications to make.
        filename_hint: Base name for the modified file (UUID suffix added automatically).
    """

    # 0. Download the original file and detect its type
    try:
        input_file_path, file_type = await download_file(file_id)
    except RuntimeError as e:
        raise ToolError(f"Failed to download file: {e}")

    config = FILE_TYPE_CONFIG[file_type]

    if "modify_system_prompt" not in config:
        raise ToolError(
            f"File type '{config['label']}' is not supported for modification. "
            f"Use analyze_file instead for CSV data."
        )

    # 1. Setup output path
    filename = f"{filename_hint}_{uuid.uuid4().hex[:8]}{config['ext']}"
    file_path = os.path.join(STORAGE_DIR, filename)

    # 2. The Healing Loop
    max_attempts = 4
    current_code = ""
    last_error = ""

    log.info(f"=== New {config['label']} modify request: '{instructions}' | input: {input_file_path} | output: {filename} ===")

    try:
        for attempt in range(max_attempts):
            try:
                # A. Generate Code (Fresh or Fix)
                system_prompt = config["modify_system_prompt"]
                if attempt == 0:
                    log.info(f"--- Attempt {attempt + 1}/{max_attempts}: fresh generation ---")
                    current_code = await generate_code(
                        instructions, system_prompt,
                        input_file_path=input_file_path,
                    )
                else:
                    log.info(f"--- Attempt {attempt + 1}/{max_attempts}: self-correction (error: {last_error}) ---")
                    current_code = await generate_code(
                        instructions, system_prompt,
                        last_error, current_code,
                        input_file_path=input_file_path,
                    )

                # B. Execute in isolated subprocess
                await run_code_in_subprocess(current_code, file_path, input_file_path=input_file_path)

                # C. Verify File Creation
                if not os.path.exists(file_path):
                    raise Exception(
                        f"Code ran without error, but no file was found at '{file_path}'. "
                        "Did you forget to save the modified file?"
                    )

                # D. Upload to Open WebUI
                log.info(f"Uploading {filename} to Open WebUI...")
                download_url = await upload_file(file_path, filename)

                log.info(f"Success on attempt {attempt + 1}: {download_url}")
                return (
                    f"Modified {config['label']} file generated successfully "
                    f"(attempt {attempt + 1} of {max_attempts}).\n"
                    f"Download: {download_url}"
                )

            except ToolError:
                raise
            except Exception as e:
                last_error = str(e)
                log.error(f"Attempt {attempt + 1} failed: {last_error}", exc_info=True)
                # Clean up partial output file before retrying
                try:
                    os.remove(file_path)
                except OSError:
                    pass

        # 3. Failure after all attempts
        log.error(f"All {max_attempts} attempts exhausted. Last error: {last_error}")
        raise ToolError(
            f"Failed to modify {config['label']} file after {max_attempts} attempts. "
            f"Last error: {last_error}"
        )
    finally:
        # Clean up the downloaded input file
        try:
            os.remove(input_file_path)
        except OSError:
            pass


@mcp.tool
async def analyze_file(
    file_id: str,
    instructions: str,
) -> str:
    """Analyze a CSV or Excel file using natural language instructions.

    Downloads the file from Open WebUI by file_id, runs an agentic loop
    where an LLM generates and executes pandas analysis code across multiple
    rounds, then a checker sub-agent verifies the results. Returns natural
    language analysis results followed by an audit trail of steps and code.

    Args:
        file_id: The Open WebUI file ID of the file to analyze.
        instructions: Natural language description of the analysis to perform.
    """

    # 0. Download the file and validate type
    try:
        input_file_path, file_type = await download_file(file_id)
    except RuntimeError as e:
        raise ToolError(f"Failed to download file: {e}")

    if file_type not in ("csv", "excel"):
        # Clean up before raising
        try:
            os.remove(input_file_path)
        except OSError:
            pass
        raise ToolError(
            f"File type '{FILE_TYPE_CONFIG.get(file_type, {}).get('label', file_type)}' "
            f"is not supported for analysis. Only CSV and Excel files are supported."
        )

    log.info(f"=== New analysis request: '{instructions}' | file: {input_file_path} (type: {file_type}) ===")

    max_rounds = 3
    max_error_retries = 2
    steps: list[dict[str, Any]] = []  # {"code": ..., "output": ..., "duration_s": ...}
    analysis_start = time.monotonic()
    analysis_complete_signaled = False

    try:
        # --- PRE-ANALYSIS DATA INSPECTION ---
        log.info("Running pre-analysis data inspection...")
        inspection = await _inspect_file(input_file_path, file_type)
        log.info(f"Inspection complete ({len(inspection)} chars)")
        steps.append({"code": "(automatic data inspection)", "output": inspection, "duration_s": 0})

        # --- AGENTIC ANALYSIS LOOP ---
        for round_num in range(max_rounds):
            log.info(f"--- Analysis round {round_num + 1}/{max_rounds} ---")

            if round_num == 0:
                # First round: fresh code generation with inspection context
                messages = [
                    {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
                    {"role": "user", "content": (
                        f"Analyze this file according to these instructions: {instructions}\n\n"
                        f"DATA INSPECTION:\n{inspection}"
                    )},
                ]
            else:
                # Subsequent rounds: include prior steps, ask LLM to continue or stop
                steps_text = ""
                for i, step in enumerate(steps[1:], 1):  # skip inspection step
                    steps_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

                messages = [
                    {"role": "system", "content": ANALYSIS_FOLLOWUP_SYSTEM_PROMPT},
                    {"role": "user", "content": (
                        f"Original instructions: {instructions}\n\n"
                        f"Here are the analysis steps completed so far:\n{steps_text}\n\n"
                        f"Should more analysis be done, or is this complete?"
                    )},
                ]

            # Generate code (or completion signal)
            log.info(f"Calling LLM for analysis round {round_num + 1}...")
            raw = await _llm_call(messages)
            log.debug(f"LLM response (round {round_num + 1}):\n{raw}")

            # Check for completion signal
            if "__ANALYSIS_COMPLETE__" in raw:
                log.info(f"LLM signaled analysis complete at round {round_num + 1}")
                analysis_complete_signaled = True
                break

            code = extract_code(raw)

            # Execute with error retries
            output = None
            last_error = ""
            for retry in range(max_error_retries + 1):
                try:
                    t0 = time.monotonic()
                    output = await run_analysis_code(code, input_file_path)
                    break
                except RuntimeError as e:
                    last_error = str(e)
                    log.warning(f"Analysis code failed (retry {retry + 1}): {last_error}")
                    if retry < max_error_retries:
                        # Ask LLM to fix the code
                        fix_messages = [
                            {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
                            {"role": "user", "content": (
                                f"The previous analysis code failed.\n\n"
                                f"USER INSTRUCTIONS: {instructions}\n\n"
                                f"DATA INSPECTION:\n{inspection}\n\n"
                                f"FAILED CODE:\n{code}\n\n"
                                f"ERROR MESSAGE:\n{last_error}\n\n"
                                f"Rewrite the code to fix this error. "
                                f"Ensure you still use `input_file_path`."
                            )},
                        ]
                        raw_fix = await _llm_call(fix_messages)
                        code = extract_code(raw_fix)

            if output is None:
                log.error(f"Round {round_num + 1}: all error retries exhausted. Last error: {last_error}")
                raise ToolError(
                    f"Analysis code failed after {max_error_retries + 1} attempts. "
                    f"Last error: {last_error}"
                )

            duration = time.monotonic() - t0
            steps.append({"code": code, "output": output, "duration_s": round(duration, 1)})
            log.info(f"Round {round_num + 1} complete. Output length: {len(output)} chars, duration: {duration:.1f}s")

        # Check for analysis content beyond inspection
        analysis_steps = [s for s in steps if s["code"] != "(automatic data inspection)"]
        if not analysis_steps:
            raise ToolError("Analysis produced no results. The LLM did not generate any code.")

        if not analysis_complete_signaled:
            log.warning(f"Analysis reached maximum rounds ({max_rounds}) without explicit completion signal")

        # --- SUB-AGENT CHECKER ---
        checker_status, analysis_result = await _run_checker(instructions, analysis_steps)
        checker_note = "PASSED"
        checker_caveat = ""

        # If checker failed, redo once with feedback
        if checker_status == "failed":
            log.info(f"Checker flagged issues, running one redo. Feedback: {analysis_result[:200]}")
            checker_feedback = analysis_result

            redo_steps = await _run_analysis_redo(instructions, checker_feedback, input_file_path, steps)
            steps.extend(redo_steps)

            # Re-check (accept whatever comes out)
            redo_status, analysis_result = await _run_checker(instructions, [s for s in steps if s["code"] != "(automatic data inspection)"])
            if redo_status == "failed":
                # Accept anyway on second pass
                checker_note = "ACCEPTED WITH CAVEATS"
                checker_caveat = f"\n  - Checker feedback: {analysis_result[:300]}"
                # Use the redo result as the summary if it's not just an error message
                if not analysis_result.startswith("EXECUTION FAILED"):
                    pass  # analysis_result is already set from checker
            else:
                checker_note = "PASSED (after correction)"

        # --- FORMAT OUTPUT ---
        total_time = time.monotonic() - analysis_start
        analysis_rounds = len([s for s in steps if s["code"] != "(automatic data inspection)"])
        redo_happened = checker_note != "PASSED"

        qa_section = (
            f"## Quality Assurance\n"
            f"- Checker: {checker_note}{checker_caveat}\n"
            f"- Analysis rounds: {analysis_rounds}{' (+ redo)' if redo_happened else ''}\n"
            f"- Total execution time: {total_time:.1f}s\n"
        )
        if not analysis_complete_signaled:
            qa_section += f"- Note: Analysis reached maximum rounds ({max_rounds}) without explicit completion signal\n"

        methodology = ""
        for i, step in enumerate(steps, 0):
            if step["code"] == "(automatic data inspection)":
                methodology += f"\n### Data Inspection\n**Output:**\n```\n{step['output']}\n```\n"
            else:
                duration_note = f" ({step.get('duration_s', '?')}s)" if step.get("duration_s") else ""
                methodology += (
                    f"\n### Step {i}{duration_note}\n"
                    f"**Code:**\n```python\n{step['code']}\n```\n"
                    f"**Output:**\n```\n{step['output']}\n```\n"
                )

        return (
            f"## Analysis Results\n\n{analysis_result}\n\n"
            f"---\n\n{qa_section}\n"
            f"---\n\n## Methodology\n{methodology}"
        )

    finally:
        try:
            os.remove(input_file_path)
        except OSError:
            pass


async def _run_checker(instructions: str, steps: list[dict[str, str]]) -> tuple[str, str]:
    """Run the sub-agent checker on analysis results.

    Returns (status, content) where status is 'passed', 'failed', or 'unknown'.
    """
    steps_text = ""
    for i, step in enumerate(steps, 1):
        steps_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

    messages = [
        {"role": "system", "content": ANALYSIS_CHECKER_SYSTEM_PROMPT},
        {"role": "user", "content": (
            f"Original instructions: {instructions}\n\n"
            f"Analysis steps:\n{steps_text}"
        )},
    ]
    log.info("Running analysis checker sub-agent...")
    raw = await _llm_call(messages)
    status, content = _parse_checker_result(raw)
    log.info(f"Checker status: {status}")
    log.debug(f"Checker content:\n{content[:500]}")
    return status, content


async def _run_analysis_redo(
    instructions: str,
    checker_feedback: str,
    input_file_path: str,
    prior_steps: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Re-run analysis with checker feedback. Returns new steps.

    On execution failure, records the failed attempt in the returned steps
    (with error message as output) instead of silently dropping it.
    """
    prior_text = ""
    for i, step in enumerate(prior_steps, 1):
        prior_text += f"\n--- Step {i} Code ---\n{step['code']}\n--- Step {i} Output ---\n{_truncate_output(step['output'])}\n"

    messages = [
        {"role": "system", "content": ANALYSIS_SYSTEM_PROMPT},
        {"role": "user", "content": (
            f"Analyze this file according to these instructions: {instructions}\n\n"
            f"A reviewer found these issues with the previous analysis:\n{checker_feedback}\n\n"
            f"Previous analysis steps for context:\n{prior_text}\n\n"
            f"Write corrected analysis code that addresses the reviewer's feedback."
        )},
    ]
    log.info("Running analysis redo based on checker feedback...")
    raw = await _llm_call(messages)
    code = extract_code(raw)

    new_steps = []
    try:
        t0 = time.monotonic()
        output = await run_analysis_code(code, input_file_path)
        duration = time.monotonic() - t0
        new_steps.append({"code": code, "output": output, "duration_s": round(duration, 1)})
    except RuntimeError as e:
        log.error(f"Redo analysis code failed: {e}")
        new_steps.append({"code": code, "output": f"EXECUTION FAILED: {e}", "duration_s": 0})

    return new_steps


@mcp.tool
async def conduct_deep_research(
    instructions: str,
) -> str:
    """Conduct deep research on a topic using web search and synthesis.

    Uses the OpenAI Responses API with the o4-mini-deep-research model to
    search and synthesize hundreds of web sources into a comprehensive
    research report. This can take several minutes to complete.

    Args:
        instructions: Natural language description of the research to conduct.
    """
    log.info(f"=== New deep research request: '{instructions[:100]}' ===")
    t0 = time.monotonic()

    # 1. Create background research task
    try:
        response = await client.responses.create(
            model=DEEP_RESEARCH_MODEL,
            input=instructions,
            background=True,
            tools=[{"type": "web_search_preview"}],
        )
    except (APIError, APIConnectionError, RateLimitError) as e:
        log.error(f"Deep research API call failed: {e}")
        raise ToolError(f"Failed to start deep research: {e}")

    response_id = response.id
    log.info(f"Deep research created: id={response_id}, status={response.status}")

    # 2. Poll for completion
    last_status = response.status
    while response.status in ("queued", "in_progress"):
        if time.monotonic() - t0 > DEEP_RESEARCH_TIMEOUT:
            log.error(f"Deep research timed out after {DEEP_RESEARCH_TIMEOUT}s (id={response_id})")
            raise ToolError(
                f"Deep research timed out after {DEEP_RESEARCH_TIMEOUT} seconds. "
                f"The research may still be running — try again later."
            )

        await asyncio.sleep(DEEP_RESEARCH_POLL_INTERVAL)

        try:
            response = await client.responses.retrieve(response_id)
        except (APIError, APIConnectionError, RateLimitError) as e:
            log.warning(f"Poll failed (will retry): {e}")
            continue

        if response.status != last_status:
            elapsed = time.monotonic() - t0
            log.info(f"Deep research status: {last_status} -> {response.status} ({elapsed:.0f}s elapsed)")
            last_status = response.status

    elapsed = time.monotonic() - t0
    log.info(f"Deep research finished: status={response.status}, elapsed={elapsed:.1f}s")

    # 3. Handle terminal status
    if response.status != "completed":
        # Extract error details from the response
        error_detail = ""
        try:
            # The API may include error info in various places
            if hasattr(response, "error") and response.error:
                error_detail = str(response.error)
            elif hasattr(response, "last_error") and response.last_error:
                error_detail = str(response.last_error)
            elif hasattr(response, "incomplete_details") and response.incomplete_details:
                error_detail = str(response.incomplete_details)
            # Also dump output items for clues
            if hasattr(response, "output") and response.output:
                for item in response.output:
                    log.info(f"Response output item: type={getattr(item, 'type', '?')}, content={str(item)[:300]}")
        except Exception:
            pass
        log.error(f"Deep research failed: status={response.status}, error={error_detail}, response_id={response_id}")
        log.debug(f"Full failed response: {response}")
        raise ToolError(
            f"Deep research ended with status '{response.status}' after {elapsed:.0f}s. "
            f"{f'Error: {error_detail}. ' if error_detail else ''}"
            f"Please try again."
        )

    # 4. Extract results
    report = response.output_text or ""
    if not report:
        raise ToolError("Deep research completed but returned no output.")

    # 5. Extract source citations from annotations
    sources = []
    try:
        for item in response.output:
            if hasattr(item, "content"):
                for block in item.content:
                    if hasattr(block, "annotations"):
                        for ann in block.annotations:
                            url = getattr(ann, "url", None)
                            title = getattr(ann, "title", None)
                            if url and url not in [s["url"] for s in sources]:
                                sources.append({"url": url, "title": title or url})
    except Exception:
        log.debug("Could not extract annotations from response", exc_info=True)

    # 6. Format output
    result = f"## Deep Research Report\n\n{report}"

    if sources:
        result += "\n\n---\n\n## Sources\n"
        for src in sources:
            result += f"- [{src['title']}]({src['url']})\n"

    result += f"\n---\n*Completed in {elapsed:.0f}s*"

    log.info(f"Deep research complete: {len(report)} chars, {len(sources)} sources, {elapsed:.1f}s")
    return result


# --- HEALTH CHECK ---

@mcp.custom_route("/health", methods=["GET"])
async def health(request: Request) -> JSONResponse:
    """Health check endpoint for orchestrators and monitoring."""
    return JSONResponse({"status": "ok"})

# --- ASGI APP ---

app = mcp.http_app(path="/mcp", stateless_http=True)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)